"""Data export functionality for Excel/CSV downloads"""
import csv
import io
from datetime import datetime
from django.http import HttpResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from apps.cases.models import OpcRegistration, OpcVisit
from apps.inventory.models import StockLevel, StockMovement
from apps.facilities.models import Facility


def _get_accessible_cases(request):
    """Get cases accessible to the current user"""
    accessible = request.user.get_accessible_facilities()
    qs = OpcRegistration.objects.all().select_related('facility', 'registered_by')
    if accessible is not None:
        qs = qs.filter(facility__in=accessible)
    return qs


def _style_excel_header(ws, headers):
    """Apply styling to Excel header row"""
    header_fill = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_cases_excel(request):
    """Export cases to Excel"""
    format_type = request.query_params.get('format', 'excel')  # excel or csv
    case_type = request.query_params.get('type', 'all')  # SAM, MAM, or all
    status = request.query_params.get('status', 'all')
    
    qs = _get_accessible_cases(request)
    
    if case_type != 'all':
        qs = qs.filter(malnutrition_type=case_type.upper())
    if status != 'all':
        qs = qs.filter(status=status)
    
    if format_type == 'csv':
        return _export_cases_csv(qs)
    return _export_cases_excel(qs)


def _export_cases_excel(qs):
    """Generate Excel file for cases"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Cases"
    
    headers = [
        'Registration No', 'Child Name', 'Registration Date', 'Date of Birth', 'Age (months)',
        'Gender', 'Caregiver Name', 'Caregiver Phone', 'Caregiver Relationship',
        'Malnutrition Type', 'Status', 'Facility', 'Region', 'District',
        'Admission Date', 'Admission Type', 'MAM Type',
        'Admission Weight (kg)',
        'Current Weight (kg)', 'Height (cm)', 'MUAC (cm)', 'Oedema',
        'Discharge Date', 'Outcome', 'Notes'
    ]
    
    _style_excel_header(ws, headers)
    
    for idx, case in enumerate(qs, 2):
        visit = OpcVisit.objects.filter(registration=case).order_by('-visit_date').first()
        
        ws.cell(row=idx, column=1, value=case.registration_number or case.id)
        ws.cell(row=idx, column=2, value=case.child_name)
        ws.cell(row=idx, column=3, value=case.registration_date.strftime('%Y-%m-%d') if case.registration_date else '')
        ws.cell(row=idx, column=4, value=case.date_of_birth.strftime('%Y-%m-%d') if case.date_of_birth else '')
        ws.cell(row=idx, column=5, value=case.age_months)
        ws.cell(row=idx, column=6, value=case.child_gender)
        ws.cell(row=idx, column=7, value=case.caregiver_name)
        ws.cell(row=idx, column=8, value=case.caregiver_phone or '')
        ws.cell(row=idx, column=9, value=case.caregiver_relationship or '')
        ws.cell(row=idx, column=10, value=case.malnutrition_type)
        ws.cell(row=idx, column=11, value=case.status)
        ws.cell(row=idx, column=12, value=case.facility.name if case.facility else '')
        ws.cell(row=idx, column=13, value=case.facility.district.region.name if case.facility and case.facility.district and case.facility.district.region else '')
        ws.cell(row=idx, column=14, value=case.facility.district.name if case.facility and case.facility.district else '')
        ws.cell(row=idx, column=15, value=case.admission_date.strftime('%Y-%m-%d') if case.admission_date else '')
        ws.cell(row=idx, column=16, value=case.admission_type or '')
        ws.cell(row=idx, column=17, value=case.mam_type or '')
        ws.cell(row=idx, column=18, value=float(case.weight_kg) if case.weight_kg else '')
        ws.cell(row=idx, column=19, value=float(visit.weight_kg) if visit else float(case.weight_kg) if case.weight_kg else '')
        ws.cell(row=idx, column=20, value=float(visit.height_cm) if visit else float(case.height_cm) if case.height_cm else '')
        ws.cell(row=idx, column=21, value=float(visit.muac_cm) if visit and visit.muac_cm else float(case.muac_cm) if case.muac_cm else '')
        ws.cell(row=idx, column=22, value=case.oedema or 'No')
        ws.cell(row=idx, column=23, value=case.discharge_date.strftime('%Y-%m-%d') if case.discharge_date else '')
        ws.cell(row=idx, column=24, value=case.outcome or '')
        ws.cell(row=idx, column=25, value=case.outcome_notes or '')
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except (TypeError, AttributeError):
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="cases_{timestamp}.xlsx"'
    return response


def _export_cases_csv(qs):
    """Generate CSV file for cases"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    headers = [
        'Registration No', 'Child Name', 'Registration Date', 'Date of Birth', 'Age (months)',
        'Gender', 'Caregiver Name', 'Caregiver Phone', 'Caregiver Relationship',
        'Malnutrition Type', 'Status', 'Facility', 'Region', 'District',
        'Admission Date', 'Admission Type', 'MAM Type',
        'Admission Weight (kg)', 'Current Weight (kg)',
        'Height (cm)', 'MUAC (cm)', 'Oedema', 'Discharge Date', 'Outcome', 'Notes'
    ]
    writer.writerow(headers)
    
    for case in qs:
        visit = OpcVisit.objects.filter(registration=case).order_by('-visit_date').first()
        writer.writerow([
            case.registration_number or case.id,
            case.child_name,
            case.registration_date.strftime('%Y-%m-%d') if case.registration_date else '',
            case.date_of_birth.strftime('%Y-%m-%d') if case.date_of_birth else '',
            case.age_months,
            case.child_gender,
            case.caregiver_name,
            case.caregiver_phone or '',
            case.caregiver_relationship or '',
            case.malnutrition_type,
            case.status,
            case.facility.name if case.facility else '',
            case.facility.district.region.name if case.facility and case.facility.district and case.facility.district.region else '',
            case.facility.district.name if case.facility and case.facility.district else '',
            case.admission_date.strftime('%Y-%m-%d') if case.admission_date else '',
            case.admission_type or '',
            case.mam_type or '',
            float(case.weight_kg) if case.weight_kg else '',
            float(visit.weight_kg) if visit else '',
            float(visit.height_cm) if visit and visit.height_cm else float(case.height_cm) if case.height_cm else '',
            float(visit.muac_cm) if visit and visit.muac_cm else float(case.muac_cm) if case.muac_cm else '',
            case.oedema or 'No',
            case.discharge_date.strftime('%Y-%m-%d') if case.discharge_date else '',
            case.outcome or '',
            case.outcome_notes or '',
        ])
    
    response = HttpResponse(output.getvalue(), content_type='text/csv')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="cases_{timestamp}.csv"'
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_inventory_excel(request):
    """Export inventory data to Excel/CSV"""
    format_type = request.query_params.get('format', 'excel')
    facility_id = request.query_params.get('facility')
    
    qs = StockLevel.objects.all().select_related('facility', 'item', 'item__category')
    
    accessible = request.user.get_accessible_facilities()
    if accessible is not None:
        qs = qs.filter(facility__in=accessible)
    
    if facility_id:
        qs = qs.filter(facility_id=facility_id)
    
    if format_type == 'csv':
        return _export_inventory_csv(qs)
    return _export_inventory_excel(qs)


def _export_inventory_excel(qs):
    """Generate Excel file for inventory"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"
    
    headers = [
        'Item Name', 'Category', 'Facility', 'Region', 'District',
        'Current Quantity', 'Unit', 'Minimum Stock', 'Reorder Level',
        'Last Updated', 'Batch Number', 'Expiry Date'
    ]
    
    _style_excel_header(ws, headers)
    
    for idx, stock in enumerate(qs, 2):
        item = stock.inventory_item
        ws.cell(row=idx, column=1, value=item.name if item else '')
        ws.cell(row=idx, column=2, value=item.category if item else '')
        ws.cell(row=idx, column=3, value=stock.facility.name if stock.facility else '')
        ws.cell(row=idx, column=4, value=stock.facility.district.region.name if stock.facility and stock.facility.district and stock.facility.district.region else '')
        ws.cell(row=idx, column=5, value=stock.facility.district.name if stock.facility and stock.facility.district else '')
        ws.cell(row=idx, column=6, value=stock.current_stock)
        ws.cell(row=idx, column=7, value=item.unit_of_measure if item else '')
        ws.cell(row=idx, column=8, value=item.min_stock_level if item else '')
        ws.cell(row=idx, column=9, value=item.reorder_level if item else '')
        ws.cell(row=idx, column=10, value=stock.last_updated.strftime('%Y-%m-%d %H:%M') if stock.last_updated else '')
        ws.cell(row=idx, column=11, value=item.batch_number or '')
        ws.cell(row=idx, column=12, value=item.expiry_date.strftime('%Y-%m-%d') if item and item.expiry_date else '')
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except (TypeError, AttributeError):
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="inventory_{timestamp}.xlsx"'
    return response


def _export_inventory_csv(qs):
    """Generate CSV file for inventory"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    headers = [
        'Item Name', 'Category', 'Facility', 'Current Quantity', 'Unit',
        'Minimum Stock', 'Reorder Level', 'Last Updated', 'Batch Number', 'Expiry Date'
    ]
    writer.writerow(headers)
    
    for stock in qs:
        item = stock.inventory_item
        writer.writerow([
            item.name if item else '',
            item.category if item else '',
            stock.facility.name if stock.facility else '',
            stock.current_stock,
            item.unit_of_measure if item else '',
            item.min_stock_level if item else '',
            item.reorder_level if item else '',
            stock.last_updated.strftime('%Y-%m-%d %H:%M') if stock.last_updated else '',
            item.batch_number or '',
            item.expiry_date.strftime('%Y-%m-%d') if item and item.expiry_date else '',
        ])
    
    response = HttpResponse(output.getvalue(), content_type='text/csv')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="inventory_{timestamp}.csv"'
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_options(request):
    """Get available export options/filters"""
    accessible = request.user.get_accessible_facilities()
    
    facilities = []
    if accessible is not None:
        facilities = list(accessible.values('id', 'name'))
    else:
        facilities = list(Facility.objects.all().values('id', 'name'))
    
    return Response({
        'success': True,
        'data': {
            'case_types': ['SAM', 'MAM', 'all'],
            'case_statuses': ['Active', 'Discharged', 'Defaulted', 'all'],
            'facilities': facilities,
            'formats': ['excel', 'csv']
        }
    })
