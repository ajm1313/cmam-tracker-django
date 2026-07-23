"""Data import functionality for Excel/CSV uploads"""
import csv
import io
from datetime import datetime
from django.db import transaction
from django.http import JsonResponse, HttpResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill
from apps.cases.models import OpcRegistration
from apps.inventory.models import InventoryItem, StockLevel
from apps.facilities.models import Facility


def validate_required_fields(data, required_fields):
    """Validate that all required fields are present"""
    missing = [f for f in required_fields if not data.get(f)]
    return missing


def parse_date(date_str):
    """Parse date string in various formats"""
    if not date_str:
        return None
    formats = ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y', '%Y/%m/%d']
    for fmt in formats:
        try:
            return datetime.strptime(str(date_str).strip(), fmt).date()
        except ValueError:
            continue
    return None


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_cases_preview(request):
    """Preview case import data before actual import"""
    file_obj = request.FILES.get('file')
    facility_id = request.data.get('facility_id')
    malnutrition_type = request.data.get('malnutrition_type')
    
    if not file_obj:
        return Response({'success': False, 'error': 'No file provided'}, status=400)
    
    file_ext = file_obj.name.split('.')[-1].lower()
    
    try:
        if file_ext == 'csv':
            preview_data = _preview_cases_csv(file_obj, facility_id, malnutrition_type)
        elif file_ext in ['xlsx', 'xls']:
            preview_data = _preview_cases_excel(file_obj, facility_id, malnutrition_type)
        else:
            return Response({'success': False, 'error': 'Unsupported file format. Use CSV or Excel.'}, status=400)
        
        return Response({
            'success': True,
            'data': {
                'total_rows': preview_data['total'],
                'valid_rows': preview_data['valid'],
                'invalid_rows': preview_data['invalid'],
                'preview': preview_data['rows'][:10],  # First 10 rows only
                'errors': preview_data['errors'][:10]
            }
        })
    except Exception as e:
        return Response({'success': False, 'error': str(e)}, status=400)


def _preview_cases_csv(file_obj, facility_id=None, malnutrition_type=None):
    """Preview CSV cases data"""
    content = file_obj.read().decode('utf-8')
    reader = csv.DictReader(io.StringIO(content))
    return _process_case_preview(reader, facility_id, malnutrition_type)


def _preview_cases_excel(file_obj, facility_id=None, malnutrition_type=None):
    """Preview Excel cases data"""
    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        rows.append(row_dict)
    
    return _process_case_preview(rows, facility_id, malnutrition_type)


def _process_case_preview(rows, default_facility_id=None, default_malnutrition_type=None):
    """Process and validate case import preview"""
    results = []
    errors = []
    valid_count = 0
    
    # If a default facility is provided, facility_id is not required per-row
    if default_facility_id:
        required = ['child_name', 'child_gender', 'date_of_birth']
    else:
        required = ['child_name', 'child_gender', 'date_of_birth', 'facility_id']
    
    for idx, row in enumerate(rows, 2):
        row_data = {k: v for k, v in row.items() if k}
        
        # Apply default facility if provided and row doesn't have one
        if default_facility_id and not row_data.get('facility_id'):
            row_data['facility_id'] = default_facility_id
        
        # Apply default malnutrition type if provided and row doesn't have one
        if default_malnutrition_type and not row_data.get('malnutrition_type'):
            row_data['malnutrition_type'] = default_malnutrition_type
        
        # Determine case type for type-specific validation
        mal_type = str(row_data.get('malnutrition_type', 'SAM')).strip().upper()
        is_ipc = mal_type == 'IPC'
        
        # IPC cases don't require date_of_birth (they use age_months)
        row_required = list(required)
        if is_ipc and 'date_of_birth' in row_required:
            row_required.remove('date_of_birth')
        if is_ipc and 'age_months' not in row_required:
            row_required.append('age_months')
        
        # Validate required fields
        missing = validate_required_fields(row_data, row_required)
        
        # Validate facility exists
        facility_error = None
        if row_data.get('facility_id'):
            try:
                Facility.objects.get(id=row_data['facility_id'])
            except Facility.DoesNotExist:
                facility_error = f"Facility ID {row_data['facility_id']} not found"
        
        # Validate gender
        gender_error = None
        if row_data.get('child_gender') and row_data['child_gender'] not in ['Male', 'Female']:
            gender_error = f"Gender must be 'Male' or 'Female'"
        
        # Validate date
        date_error = None
        if row_data.get('date_of_birth'):
            parsed = parse_date(row_data['date_of_birth'])
            if not parsed:
                date_error = f"Invalid date format for date_of_birth"
        
        is_valid = not missing and not facility_error and not gender_error and not date_error
        if is_valid:
            valid_count += 1
        
        row_errors = []
        if missing:
            row_errors.append(f"Missing required: {', '.join(missing)}")
        if facility_error:
            row_errors.append(facility_error)
        if gender_error:
            row_errors.append(gender_error)
        if date_error:
            row_errors.append(date_error)
        
        if row_errors:
            errors.append({'row': idx, 'errors': row_errors})
        
        results.append({
            'row': idx,
            'data': {
                'child_name': row_data.get('child_name', '')[:50],
                'child_gender': row_data.get('child_gender', ''),
                'date_of_birth': str(row_data.get('date_of_birth', ''))[:20],
                'facility_id': row_data.get('facility_id', ''),
                'malnutrition_type': row_data.get('malnutrition_type', 'SAM'),
            },
            'valid': is_valid
        })
    
    return {
        'total': len(results),
        'valid': valid_count,
        'invalid': len(results) - valid_count,
        'rows': results,
        'errors': errors
    }


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_cases_execute(request):
    """Execute case import from preview"""
    file_obj = request.FILES.get('file')
    facility_id = request.data.get('facility_id')
    malnutrition_type = request.data.get('malnutrition_type')
    if not file_obj:
        return Response({'success': False, 'error': 'No file provided'}, status=400)
    
    file_ext = file_obj.name.split('.')[-1].lower()
    
    try:
        if file_ext == 'csv':
            result = _import_cases_csv(file_obj, request.user, facility_id, malnutrition_type)
        elif file_ext in ['xlsx', 'xls']:
            result = _import_cases_excel(file_obj, request.user, facility_id, malnutrition_type)
        else:
            return Response({'success': False, 'error': 'Unsupported file format'}, status=400)
        
        return Response({
            'success': True,
            'data': result
        })
    except Exception as e:
        return Response({'success': False, 'error': str(e)}, status=400)


def _import_cases_csv(file_obj, user, default_facility_id=None, default_malnutrition_type=None):
    """Import cases from CSV"""
    content = file_obj.read().decode('utf-8')
    reader = csv.DictReader(io.StringIO(content))
    return _execute_case_import(reader, user, default_facility_id, default_malnutrition_type)


def _import_cases_excel(file_obj, user, default_facility_id=None, default_malnutrition_type=None):
    """Import cases from Excel"""
    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        rows.append(row_dict)
    
    return _execute_case_import(rows, user, default_facility_id, default_malnutrition_type)


@transaction.atomic
def _execute_case_import(rows, user, default_facility_id=None, default_malnutrition_type=None):
    """Execute the actual case import"""
    created = 0
    failed = 0
    errors = []
    
    accessible_facilities = user.get_accessible_facilities()
    if accessible_facilities is not None:
        accessible_ids = set(accessible_facilities.values_list('id', flat=True))
    else:
        accessible_ids = None
    
    for idx, row in enumerate(rows, 2):
        try:
            row_data = {k: v for k, v in row.items() if k}
            
            # Apply default facility if provided and row doesn't have one
            if default_facility_id and not row_data.get('facility_id'):
                row_data['facility_id'] = default_facility_id
            
            # Apply default malnutrition type if provided and row doesn't have one
            if default_malnutrition_type and not row_data.get('malnutrition_type'):
                row_data['malnutrition_type'] = default_malnutrition_type
            
            # Check facility access
            facility_id = row_data.get('facility_id')
            if not facility_id:
                failed += 1
                errors.append(f"Row {idx}: No facility_id provided")
                continue
            if accessible_ids is not None and int(facility_id) not in accessible_ids:
                failed += 1
                errors.append(f"Row {idx}: No access to facility {facility_id}")
                continue
            
            # Get facility
            facility = Facility.objects.get(id=facility_id)
            
            # Determine malnutrition type
            mal_type = str(row_data.get('malnutrition_type', 'SAM')).strip().upper()
            
            # Parse dates
            dob = parse_date(row_data.get('date_of_birth'))
            reg_date = parse_date(row_data.get('registration_date')) or datetime.now().date()
            
            # Create case
            case = OpcRegistration.objects.create(
                child_name=row_data.get('child_name', '').strip(),
                child_gender=row_data.get('child_gender', 'Male'),
                date_of_birth=dob,
                age_months=int(row_data.get('age_months', 0)) if row_data.get('age_months') else 0,
                facility=facility,
                registration_date=reg_date,
                admission_date=parse_date(row_data.get('admission_date')) or reg_date,
                malnutrition_type=mal_type,
                status='Active',
                caregiver_name=row_data.get('caregiver_name', '') or row_data.get('guardian_name', '').strip() or 'Unknown',
                caregiver_phone=(row_data.get('caregiver_phone', '') or row_data.get('guardian_phone', '')).strip()[:20] or '',
                caregiver_relationship=row_data.get('caregiver_relationship', '') or None,
                address=row_data.get('community', '') or row_data.get('address', '') or None,
                weight_kg=float(row_data.get('weight_kg', 0)) if row_data.get('weight_kg') else 0,
                muac_cm=float(row_data.get('muac_cm', 0)) if row_data.get('muac_cm') else None,
                height_cm=float(row_data.get('height_cm', 0)) if row_data.get('height_cm') else 0,
                oedema=row_data.get('oedema', '') or None,
                # Extended fields
                mam_type=row_data.get('mam_type', '') or None,
                admission_type=row_data.get('admission_type', 'New Admission') or 'New Admission',
                referral_source=row_data.get('referral_source', '') or None,
                z_score_wfh=row_data.get('z_score_wfh', '') or None,
                z_score_wfa=row_data.get('z_score_wfa', '') or None,
                z_score_hfa=row_data.get('z_score_hfa', '') or None,
                appetite_test=row_data.get('appetite_test', '') or None,
                medical_complications=str(row_data.get('medical_complications', '')).lower() in ('yes', 'true', '1'),
                complications_notes=row_data.get('complications_notes', '') or None,
                house_location=row_data.get('house_location', '') or None,
                travel_time=row_data.get('travel_time', '') or None,
                father_alive=row_data.get('father_alive', '') or None,
                mother_alive=row_data.get('mother_alive', '') or None,
                diarrhoea=row_data.get('diarrhoea', '') or None,
                stool_frequency=row_data.get('stool_frequency', '') or None,
                vomiting=row_data.get('vomiting', '') or None,
                cough=row_data.get('cough', '') or None,
                passing_urine=row_data.get('passing_urine', '') or None,
                oedema_duration_days=int(row_data.get('oedema_duration_days', 0)) if row_data.get('oedema_duration_days') else None,
                breastfeeding_status=row_data.get('breastfeeding_status', '') or None,
                breastfeeding_prospect=row_data.get('breastfeeding_prospect', '') or None,
                immunization_status=row_data.get('immunization_status', '') or None,
                g6pd_status=row_data.get('g6pd_status', '') or None,
                respiratory_rate=row_data.get('respiratory_rate', '') or None,
                temperature_celsius=float(row_data.get('temperature_celsius', 0)) if row_data.get('temperature_celsius') else None,
                chest_indrawing=row_data.get('chest_indrawing', '') or None,
                eyes_condition=row_data.get('eyes_condition', '') or None,
                conjunctiva=row_data.get('conjunctiva', '') or None,
                ears_condition=row_data.get('ears_condition', '') or None,
                mouth_condition=row_data.get('mouth_condition', '') or None,
                lymph_nodes=row_data.get('lymph_nodes', '') or None,
                hands_feet=row_data.get('hands_feet', '') or None,
                skin_changes=row_data.get('skin_changes', '') or None,
                disability=row_data.get('disability', '') or None,
                disability_details=row_data.get('disability_details', '') or None,
                amoxicillin_date=parse_date(row_data.get('amoxicillin_date')),
                amoxicillin_dosage=row_data.get('amoxicillin_dosage', '') or None,
                vitamin_a_date=parse_date(row_data.get('vitamin_a_date')),
                vitamin_a_dosage=row_data.get('vitamin_a_dosage', '') or None,
                folic_acid_date=parse_date(row_data.get('folic_acid_date')),
                folic_acid_dosage=row_data.get('folic_acid_dosage', '') or None,
                deworming_date=parse_date(row_data.get('deworming_date')),
                deworming_dosage=row_data.get('deworming_dosage', '') or None,
                measles_vaccine_date=parse_date(row_data.get('measles_vaccine_date')),
                measles_vaccine_dosage=row_data.get('measles_vaccine_dosage', '') or None,
                malaria_test_date=parse_date(row_data.get('malaria_test_date')),
                malaria_test_result=row_data.get('malaria_test_result', '') or None,
                antimalarial_date=parse_date(row_data.get('antimalarial_date')),
                antimalarial_dosage=row_data.get('antimalarial_dosage', '') or None,
                rutf_sachets_given=int(row_data.get('rutf_sachets_given', 0)) if row_data.get('rutf_sachets_given') else None,
                rutf_ration_per_day=float(row_data.get('rutf_ration_per_day', 0)) if row_data.get('rutf_ration_per_day') else None,
                next_visit_date=parse_date(row_data.get('next_visit_date')),
                other_drug_1=row_data.get('other_drug_1', '') or None,
                other_drug_1_date=parse_date(row_data.get('other_drug_1_date')),
                other_drug_1_dosage=row_data.get('other_drug_1_dosage', '') or None,
                other_drug_2=row_data.get('other_drug_2', '') or None,
                other_drug_2_date=parse_date(row_data.get('other_drug_2_date')),
                other_drug_2_dosage=row_data.get('other_drug_2_dosage', '') or None,
                other_drug_3=row_data.get('other_drug_3', '') or None,
                other_drug_3_date=parse_date(row_data.get('other_drug_3_date')),
                other_drug_3_dosage=row_data.get('other_drug_3_dosage', '') or None,
                # MAM-specific
                previous_sam_episode=str(row_data.get('previous_sam_episode', '')).lower() in ('yes', 'true', '1'),
                failed_counselling_only=str(row_data.get('failed_counselling_only', '')).lower() in ('yes', 'true', '1'),
                hiv_tb_status=row_data.get('hiv_tb_status', '') or None,
                household_vulnerability=row_data.get('household_vulnerability', '') or None,
                poor_maternal_health=str(row_data.get('poor_maternal_health', '')).lower() in ('yes', 'true', '1'),
                mother_deceased=str(row_data.get('mother_deceased', '')).lower() in ('yes', 'true', '1'),
                food_product_type=row_data.get('food_product_type', '') or None,
                food_product_quantity=row_data.get('food_product_quantity', '') or None,
                mebendazole_date=parse_date(row_data.get('mebendazole_date')),
                counselling=row_data.get('counselling', '') or None,
                additional_notes=row_data.get('additional_notes', '') or row_data.get('notes', '') or None,
                # Additional fields
                admission_criteria=row_data.get('admission_criteria', '') or row_data.get('enrolment_criteria', '') or row_data.get('entry_criteria', '') or None,
                additional_medical_history=row_data.get('additional_medical_history', '') or None,
                physical_exam_notes=row_data.get('physical_exam_notes', '') or None,
                registration_latitude=float(row_data.get('registration_latitude', 0)) if row_data.get('registration_latitude') else None,
                registration_longitude=float(row_data.get('registration_longitude', 0)) if row_data.get('registration_longitude') else None,
                # MAM additional fields
                immunization_action=row_data.get('immunization_action', '') or None,
                other_medicines=row_data.get('other_medicines', '') or None,
                created_by=user
            )
            
            created += 1
            
        except Exception as e:
            failed += 1
            errors.append(f"Row {idx}: {str(e)}")
    
    return {
        'created': created,
        'failed': failed,
        'errors': errors[:20]  # Limit errors
    }


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_inventory_preview(request):
    """Preview inventory import data"""
    file_obj = request.FILES.get('file')
    facility_id = request.data.get('facility_id')
    
    if not file_obj:
        return Response({'success': False, 'error': 'No file provided'}, status=400)
    
    if not facility_id:
        return Response({'success': False, 'error': 'facility_id is required'}, status=400)
    
    file_ext = file_obj.name.split('.')[-1].lower()
    
    try:
        if file_ext == 'csv':
            preview_data = _preview_inventory_csv(file_obj, facility_id)
        elif file_ext in ['xlsx', 'xls']:
            preview_data = _preview_inventory_excel(file_obj, facility_id)
        else:
            return Response({'success': False, 'error': 'Unsupported file format'}, status=400)
        
        return Response({
            'success': True,
            'data': preview_data
        })
    except Exception as e:
        return Response({'success': False, 'error': str(e)}, status=400)


def _preview_inventory_csv(file_obj, facility_id):
    content = file_obj.read().decode('utf-8')
    reader = csv.DictReader(io.StringIO(content))
    return _process_inventory_preview(reader, facility_id)


def _preview_inventory_excel(file_obj, facility_id):
    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        rows.append(row_dict)
    
    return _process_inventory_preview(rows, facility_id)


def _process_inventory_preview(rows, facility_id):
    """Process and validate inventory import preview"""
    results = []
    errors = []
    valid_count = 0
    
    required = ['item_name', 'quantity']
    
    for idx, row in enumerate(rows, 2):
        row_data = {k: v for k, v in row.items() if k}
        missing = validate_required_fields(row_data, required)
        
        # Validate quantity is numeric
        qty_error = None
        if row_data.get('quantity'):
            try:
                float(row_data['quantity'])
            except ValueError:
                qty_error = "Quantity must be a number"
        
        is_valid = not missing and not qty_error
        if is_valid:
            valid_count += 1
        
        row_errors = []
        if missing:
            row_errors.append(f"Missing required: {', '.join(missing)}")
        if qty_error:
            row_errors.append(qty_error)
        
        if row_errors:
            errors.append({'row': idx, 'errors': row_errors})
        
        results.append({
            'row': idx,
            'data': {
                'item_name': str(row_data.get('item_name', ''))[:50],
                'quantity': str(row_data.get('quantity', '')),
                'batch_number': str(row_data.get('batch_number', ''))[:30],
            },
            'valid': is_valid
        })
    
    return {
        'total': len(results),
        'valid': valid_count,
        'invalid': len(results) - valid_count,
        'rows': results,
        'errors': errors[:10]
    }


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_inventory_execute(request):
    """Execute inventory import after preview"""
    file_obj = request.FILES.get('file')
    facility_id = request.data.get('facility_id')
    
    if not file_obj:
        return Response({'success': False, 'error': 'No file provided'}, status=400)
    
    if not facility_id:
        return Response({'success': False, 'error': 'facility_id is required'}, status=400)
    
    file_ext = file_obj.name.split('.')[-1].lower()
    user = request.user
    
    try:
        if file_ext == 'csv':
            result = _execute_inventory_csv(file_obj, facility_id, user)
        elif file_ext in ['xlsx', 'xls']:
            result = _execute_inventory_excel(file_obj, facility_id, user)
        else:
            return Response({'success': False, 'error': 'Unsupported file format'}, status=400)
        
        return Response({
            'success': True,
            'data': result
        })
    except Exception as e:
        return Response({'success': False, 'error': str(e)}, status=400)


def _execute_inventory_csv(file_obj, facility_id, user):
    content = file_obj.read().decode('utf-8')
    reader = csv.DictReader(io.StringIO(content))
    rows = list(reader)
    return _execute_inventory_import(rows, facility_id, user)


def _execute_inventory_excel(file_obj, facility_id, user):
    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        rows.append(row_dict)
    
    return _execute_inventory_import(rows, facility_id, user)


@transaction.atomic
def _execute_inventory_import(rows, facility_id, user):
    """Execute the actual inventory import"""
    created = 0
    updated = 0
    failed = 0
    errors = []
    
    try:
        facility = Facility.objects.get(id=facility_id)
    except Facility.DoesNotExist:
        return {'created': 0, 'updated': 0, 'failed': len(rows), 'errors': ['Facility not found']}
    
    # Check user access
    accessible_facilities = user.get_accessible_facilities()
    if accessible_facilities is not None:
        if facility not in accessible_facilities:
            return {'created': 0, 'updated': 0, 'failed': len(rows), 'errors': ['No access to this facility']}
    
    for idx, row in enumerate(rows, 2):
        try:
            row_data = {k: v for k, v in row.items() if k}
            
            item_name = str(row_data.get('item_name', '')).strip()
            if not item_name:
                failed += 1
                errors.append(f"Row {idx}: Item name is required")
                continue
            
            quantity_str = row_data.get('quantity', '0')
            try:
                quantity = float(quantity_str)
            except ValueError:
                failed += 1
                errors.append(f"Row {idx}: Invalid quantity")
                continue
            
            # Get or create inventory item
            item, item_created = InventoryItem.objects.get_or_create(
                name__iexact=item_name,
                defaults={
                    'name': item_name,
                    'code': f'IMP-{facility.code}-{idx}',
                    'category': 'Therapeutic Food',
                    'unit_of_measure': 'sachet',
                    'description': f'Imported item: {item_name}'
                }
            )
            
            # Get or create stock level
            stock_level, stock_created = StockLevel.objects.get_or_create(
                inventory_item=item,
                facility=facility,
                location_type='facility',
                defaults={'current_stock': 0}
            )
            
            # Update quantity
            stock_level.current_stock += quantity
            stock_level.save()
            
            # Create batch if batch_number provided
            batch_number = str(row_data.get('batch_number', '')).strip()
            if batch_number:
                from apps.inventory.models import ItemBatch
                expiry_str = row_data.get('expiry_date', '')
                expiry_date = parse_date(expiry_str) if expiry_str else None
                
                ItemBatch.objects.create(
                    inventory_item=item,
                    facility=facility,
                    batch_number=batch_number,
                    quantity=quantity,
                    expiry_date=expiry_date
                )
            
            # Record movement
            StockMovement.objects.create(
                inventory_item=item,
                movement_type='IN',
                quantity=quantity,
                reference_number=f"Import: {batch_number or 'N/A'}",
                destination_type='facility',
                destination_facility=facility,
                notes=f"Imported via bulk import",
                created_by=user,
                movement_date=datetime.now(),
            )
            
            if item_created:
                created += 1
            else:
                updated += 1
                
        except Exception as e:
            failed += 1
            errors.append(f"Row {idx}: {str(e)}")
    
    return {
        'created': created,
        'updated': updated,
        'failed': failed,
        'errors': errors[:20]
    }


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def import_template_download(request, model_type):
    """Download import template for cases (SAM/MAM/IPC) or inventory"""
    wb = Workbook()
    ws = wb.active
    
    if model_type == 'cases-sam':
        headers = [
            'child_name', 'child_gender', 'date_of_birth', 'age_months',
            'caregiver_name', 'caregiver_phone', 'caregiver_relationship',
            'community', 'house_location', 'travel_time',
            'father_alive', 'mother_alive',
            'referral_source', 'admission_type', 'enrolment_criteria',
            'weight_kg', 'height_cm', 'muac_cm',
            'z_score_wfh', 'z_score_wfa', 'z_score_hfa',
            'oedema', 'appetite_test',
            'medical_complications', 'complications_notes',
            'diarrhoea', 'stool_frequency', 'vomiting', 'cough', 'passing_urine',
            'oedema_duration_days', 'breastfeeding_status', 'breastfeeding_prospect',
            'immunization_status', 'g6pd_status', 'additional_medical_history',
            'respiratory_rate', 'temperature_celsius', 'chest_indrawing',
            'eyes_condition', 'conjunctiva', 'ears_condition', 'mouth_condition',
            'lymph_nodes', 'hands_feet', 'skin_changes',
            'disability', 'disability_details', 'physical_exam_notes',
            'amoxicillin_date', 'amoxicillin_dosage',
            'vitamin_a_date', 'vitamin_a_dosage',
            'folic_acid_date', 'folic_acid_dosage',
            'deworming_date', 'deworming_dosage',
            'measles_vaccine_date', 'measles_vaccine_dosage',
            'malaria_test_date', 'malaria_test_result',
            'antimalarial_date', 'antimalarial_dosage',
            'rutf_sachets_given', 'rutf_ration_per_day', 'next_visit_date',
            'other_drug_1', 'other_drug_1_date', 'other_drug_1_dosage',
            'other_drug_2', 'other_drug_2_date', 'other_drug_2_dosage',
            'other_drug_3', 'other_drug_3_date', 'other_drug_3_dosage',
            'registration_latitude', 'registration_longitude',
            'admission_date', 'registration_date', 'additional_notes'
        ]
        ws.title = "SAM Import Template"
        ws.append(headers)
        ws.append([
            'John Doe', 'Male', '2022-01-15', '24',
            'Jane Doe', '+1234567890', 'Mother',
            'Community A', 'House 12', '30 mins',
            'Yes', 'Yes',
            'Direct from community', 'New Admission', 'MUAC <11.5cm',
            '8.5', '75.0', '11.0',
            '< -3 SD', '< -3 SD', '< -3 SD',
            'None', 'Good',
            'No', '',
            'No', '', 'No', 'No', 'Yes',
            '', 'Yes', 'Good',
            'Complete for Age', 'Normal', 'No known medical history',
            '', '36.5', 'No',
            'Normal', 'Normal', 'Normal', 'Normal',
            'Normal', 'Normal', 'None',
            'No', '', 'No abnormalities noted',
            '2024-01-20', '25mg twice daily',
            '2024-01-20', '100,000 IU',
            '', '',
            '', '',
            '', '',
            '', 'Not Done',
            '', '',
            '7', '2.0', '2024-01-27',
            '', '', '',
            '', '', '',
            '', '', '',
            '6.5', '3.2',
            '2024-01-20', '2024-01-20', 'Sample SAM case'
        ])
        
    elif model_type == 'cases-mam':
        headers = [
            'child_name', 'child_gender', 'date_of_birth', 'age_months',
            'caregiver_name', 'caregiver_phone', 'caregiver_relationship',
            'community', 'house_location',
            'mam_type', 'admission_type', 'entry_criteria',
            'weight_kg', 'height_cm', 'muac_cm',
            'z_score_wfh',
            'oedema', 'appetite_test',
            'medical_complications', 'complications_notes',
            'diarrhoea', 'vomiting', 'cough',
            'breastfeeding_status', 'immunization_status', 'immunization_action',
            'previous_sam_episode', 'failed_counselling_only',
            'hiv_tb_status', 'household_vulnerability',
            'poor_maternal_health', 'mother_deceased',
            'disability', 'disability_details',
            'food_product_type', 'food_product_quantity',
            'vitamin_a_date', 'mebendazole_date', 'measles_vaccine_date',
            'other_medicines', 'counselling',
            'registration_latitude', 'registration_longitude',
            'admission_date', 'registration_date', 'additional_notes'
        ]
        ws.title = "MAM Import Template"
        ws.append(headers)
        ws.append([
            'Jane Smith', 'Female', '2022-06-10', '18',
            'Mary Smith', '+1234567890', 'Mother',
            'Community B', 'House 34',
            'High-risk MAM', 'New Admission', 'direct_new',
            '9.0', '70.0', '11.8',
            '>= -3 SD and < -2 SD',
            'None', 'Pass',
            'No', '',
            'No', 'No', 'No',
            'No', 'Complete for Age', '',
            'No', 'No',
            'None', 'None',
            'No', 'No',
            'No', '',
            'RUSF', '14 sachets',
            '2024-01-20', '', '',
            '', 'Nutrition counselling provided',
            '6.5', '3.2',
            '2024-01-20', '2024-01-20', 'Sample MAM case'
        ])
        
    elif model_type == 'cases-ipc':
        headers = [
            'child_name', 'child_gender', 'date_of_birth', 'age_months',
            'caregiver_name', 'caregiver_phone', 'caregiver_relationship',
            'house_location', 'travel_time',
            'referral_source',
            'admission_date', 'admission_type',
            'weight_kg', 'height_cm', 'muac_cm',
            'z_score_wfh',
            'oedema', 'appetite_test',
            'medical_complications', 'complications_notes',
            'temperature_celsius', 'respiratory_rate',
            'diarrhoea', 'vomiting', 'cough',
            'breastfeeding_status', 'immunization_status',
            'registration_latitude', 'registration_longitude',
            'registration_date', 'additional_notes'
        ]
        ws.title = "IPC Import Template"
        ws.append(headers)
        ws.append([
            'Baby Ali', 'Male', '2023-03-01', '15',
            'Aisha Ali', '+1234567890', 'Mother',
            'House 45, Village C', '45 mins',
            'Referred from health facility',
            '2024-01-20', 'New Admission',
            '6.5', '65.0', '10.5',
            '< -3 SD',
            '++', 'Fail',
            'Yes', 'Severe dehydration and fever',
            '38.5', '45',
            'Yes', 'Yes', 'Yes',
            'No', 'Not Complete for Age',
            '6.5', '3.2',
            '2024-01-20', 'Sample IPC case'
        ])
        
    elif model_type == 'inventory':
        headers = [
            'item_name', 'quantity', 'batch_number',
            'expiry_date', 'unit_cost', 'notes'
        ]
        ws.title = "Inventory Import Template"
        ws.append(headers)
        ws.append([
            'RUTF Sachet', '100', 'BATCH001',
            '2025-12-31', '1.50', 'Initial stock'
        ])
    else:
        return Response({'success': False, 'error': 'Invalid template type. Use: cases-sam, cases-mam, cases-ipc, or inventory'}, status=400)
    
    # Style header
    header_fill = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Auto-size columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{model_type}_import_template.xlsx"'
    return response
